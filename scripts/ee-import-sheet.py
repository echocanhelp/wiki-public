#!/usr/bin/env python3
"""One-shot: TAHSdata.xlsx → frozen _import + jsonl + sitting md + index.csv.

No U-ids in sitting markdown or the visible index. Raw jsonl keeps user_id
(gitignored). Does not write Echopedia or Google Sheets.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

VAULT = Path("/home/leedt/echo-system")
ORAL = VAULT / "knowledge" / "oral-stories"
JSONL_DIR = VAULT / "knowledge" / "interactions" / "line-stories"
DEFAULT_XLSX = ORAL / "TAHSdata-2026-08-25-operator-prune.xlsx"


def slugify(name: str) -> str:
    s = (name or "").strip().lower()
    s = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s[:60] or "unknown"


def gist(text: str, n: int = 120) -> str:
    t = re.sub(r"\s+", " ", (text or "").strip())
    if len(t) <= n:
        return t
    return t[: n - 1] + "…"


def parse_ts(v) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt)
        except ValueError:
            continue
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--export-date", default="2026-08-25")
    args = ap.parse_args()
    xlsx = Path(args.xlsx)
    if not xlsx.is_file():
        raise SystemExit(f"missing {xlsx}")

    imp = ORAL / "_import" / args.export_date
    imp.mkdir(parents=True, exist_ok=True)
    dest = imp / xlsx.name
    if xlsx.resolve() != dest.resolve():
        shutil.copy2(xlsx, dest)
    dest.chmod(0o600)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    def cell(row, name):
        i = idx.get(name)
        if i is None or i >= len(row):
            return None
        return row[i]

    JSONL_DIR.mkdir(parents=True, exist_ok=True)
    for old in JSONL_DIR.glob("*.jsonl"):
        old.unlink()

    sittings: dict[tuple[str, str], list] = defaultdict(list)
    jsonl_counts: dict[str, int] = defaultdict(int)
    n = 0
    for row in rows[1:]:
        ts = parse_ts(cell(row, "Timestamp"))
        if not ts:
            continue
        n += 1
        day = ts.strftime("%Y-%m-%d")
        display = str(cell(row, "Display Name") or "").strip() or "unknown"
        role = str(cell(row, "Role") or "").strip()
        msg = "" if cell(row, "Message") is None else str(cell(row, "Message"))
        notes = "" if cell(row, "Notes") is None else str(cell(row, "Notes"))
        lang = str(cell(row, "Language") or "").strip()
        gid = cell(row, "Group_ID")
        source = "group" if gid else "dm"
        rec = {
            "ts": ts.strftime("%Y-%m-%dT%H:%M:%S"),
            "platform": "line",
            "source_oa": "suiyue",
            "source": source,
            "display_name": display,
            "role": role,
            "text": msg,
            "notes": notes,
            "language": lang,
            "user_id": str(cell(row, "User ID") or ""),
            "group_id": str(gid or ""),
            "import_batch": args.export_date,
        }
        jpath = JSONL_DIR / f"{day}.jsonl"
        with jpath.open("a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
        jsonl_counts[day] += 1
        sittings[(display, day)].append(rec)

    index_rows = []
    written_md = []
    for (display, day), evs in sorted(sittings.items(), key=lambda x: (x[0][1], x[0][0])):
        slug = slugify(display)
        md_name = f"{day}-{slug}.md"
        md_path = ORAL / md_name
        user_texts = [
            e["text"]
            for e in evs
            if e["role"] in ("User", "Voice (transcribed)") and e["text"].strip()
        ]
        langs = [e["language"] for e in evs if e["language"]]
        lang = max(set(langs), key=langs.count) if langs else ""
        source = "group" if any(e["source"] == "group" for e in evs) else "dm"
        voice = any(e["role"] == "Voice (transcribed)" or e["notes"] == "[Voice]" for e in evs)
        user_chars = sum(len(t) for t in user_texts)
        turns = sum(1 for e in evs if e["role"] != "Bot")
        one = gist(user_texts[0] if user_texts else "")
        lines = [
            "---",
            f"date: {day}",
            f"display_name: {display}",
            f"source: {source}",
            f"lang: {lang}",
            f"turns: {turns}",
            f"user_chars: {user_chars}",
            f"voice: {str(voice).lower()}",
            "photo: false",
            "consent: unknown",
            "anonymize: false",
            f"gist: {json.dumps(one, ensure_ascii=False)}",
            "source_oa: suiyue",
            "---",
            "",
            f"# {day} · {display}",
            "",
        ]
        for e in evs:
            if e["role"] == "Bot":
                who = "Echo"
            elif e["role"] == "Voice (transcribed)":
                who = "你（語音）"
            else:
                who = "你"
            body = (e["text"] or "").rstrip()
            if not body:
                continue
            lines.append(f"**{who}** · {e['ts'][11:16]}")
            lines.append("")
            lines.append(body)
            lines.append("")
        md_path.write_text("\n".join(lines), encoding="utf-8")
        md_path.chmod(0o600)
        written_md.append(md_path)
        index_rows.append(
            {
                "date": day,
                "displayName": display,
                "source": source,
                "group_name": "",
                "lang": lang,
                "turns": turns,
                "user_chars": user_chars,
                "voice": voice,
                "photo": False,
                "consent": "unknown",
                "anonymize": False,
                "gist": one,
                "vault_path": str(md_path),
            }
        )

    idx_path = ORAL / "_index.csv"
    fields = [
        "date",
        "displayName",
        "source",
        "group_name",
        "lang",
        "turns",
        "user_chars",
        "voice",
        "photo",
        "consent",
        "anonymize",
        "gist",
        "vault_path",
    ]
    with idx_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(index_rows)
    idx_path.chmod(0o600)

    people = {d for d, _ in sittings}
    print(
        json.dumps(
            {
                "turns": n,
                "people": len(people),
                "sittings": len(sittings),
                "jsonl_days": len(jsonl_counts),
                "md_files": len(written_md),
                "import_dir": str(imp),
                "index": str(idx_path),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
